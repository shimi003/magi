from django.shortcuts import render, redirect
import warnings

# Create your views here.
from django.http import HttpResponse
from django.urls import reverse
import sdss.models as db
import json as simplejson
from django.db import models as django_models
from django.db.models import Sum
from django.db.models import Count
from django.db.models.functions import Left

from datetime import datetime, timezone, timedelta
import logging as log
# from dateutil.relativedelta import relativedelta

from sdss.BS import BSClass as b
from sdss.PL import PLClass as p
from sdss.RigidPL import RigidPLClass as r_p
from sdss.Detail import DetailClass as d
import sdss.Budget as bd
import sdss.Utility as u

GENKIN_ACC_BOT_UID = 6
DUMMY_ACC_BOT_UID = 101

#
def index(request):
    accbot_qs = db.AccBot.objects.order_by('sort_order').all()
    accbot_dic = getUidAndName(accbot_qs)
    accbot_list = getAccountList(accbot_qs)
    accbot_listgroup = getAccountListByGroup()
    context = {
        'i_list':       ['1','2','3','4','5', '6', '7', '8'],
        'b_or_c':       ['br', 'cr'],
        'journal_date': u.get_nowdt().strftime('%Y-%m-%d'),
        'accbot_dic':   accbot_dic,  # iru...?
        'account_list': accbot_list, # iru...?
        'listgroup':    accbot_listgroup,
        'view_name':    'sdss 2.0 journal input',
        'message': '',
    }
    return render(request, 'sdss.html', context)


####
def budget(request):
    context = {
        'view_name':    'sdss 2.0 budget view',
        'budget_list': getCostBudgetList(),
    }
    return render(request, 'budget.html', context)




def budget_edit_id(request, id):
    context = {
        'id':       id,
        'acc_name': db.AccBot.objects.get(uid=id).name,
        'note':     db.AccBot.objects.get(uid=id).note,
        'before':   bd.getBudget(id),
    }
    return render(request, 'budget_id.html', context)


def budget_change(request):
    # TODO 変更履歴の登録
    # なければ新規、あれば変更
    budget_qs = db.Budget.objects.filter(acc_bot_uid=db.AccBot.objects.get(uid=request.POST['id'])).order_by('for_field')
    if budget_qs.count() > 0:
        budget = budget_qs[0]
        if budget.amount_per_month != request.POST['after']:
            if budget.for_field == '' or budget.for_field > u.createYesterdayDateString():
                budget.for_field = u.createYesterdayDateString()
                budget.save()
    db.Budget.objects.create(
        acc_bot_uid=      db.AccBot.objects.get(uid=request.POST['id']),
        from_field=       u.createCurrentDateString(),
        amount_per_month= request.POST['after'],
        note=             request.POST['cause'],
    )
    return redirect('/magi/sdss/budget')


def getCostBudgetList():
    qs_bot = db.AccBot.objects.filter(acc_mid_uid__acc_top_uid=u.AccType['Cost']).order_by('sort_order')
    d = {}
    total = 0
    for entry in qs_bot:
        if not entry.acc_mid_uid.name in d:
            d[entry.acc_mid_uid.name] = []

        budget = int(bd.getBudget(entry.uid))
        note = bd.getNote(entry.uid)
        total += budget
        d[entry.acc_mid_uid.name].append({
            'uid': entry.uid,
            'name': entry.name,
            'amount': budget,
            'note': note,
        })
    d['費用予算(月間)合計'] = []
    d['費用予算(月間)合計'].append({
        'uid': '',
        'name': '',
        'amount': total,
        'note': ''
    })
    return d


def fixed_asset(request):
    context = {
        'title_jp': '固定資産一覧表示画面',
        'fixed_asset_list': getFixedAssetList(),
    }
    return render(request, 'fixed_asset.html', context)


def fixed_asset_addform(request):
    fixed_asset_acc_qs = db.AccBot.objects.filter(acc_mid_uid=3)
    fixed_asset_account_list = getAccountList(fixed_asset_acc_qs)

    context = {
        'today': u.get_nowdt().strftime('%Y-%m-%d'),
        'acc_list': fixed_asset_account_list,
        'title_jp': '固定資産登録画面',
    }
    return render(request, 'fixed_asset_add.html', context)


def fixed_asset_regist(request):
    message=''
    get_date        = request.POST['get_date']
    total           = int(request.POST['get_cost'])
    terms           = int(request.POST['amortization_term_in_month'])
    monthly_cost    = int(getCostPerMonth(total, terms))
    current_value   = int(getCurrentValue(total, monthly_cost, get_date))
    passed_month    = getPassedMonths(get_date)
    using           = True if 'is_using' in request.POST else False

    db.FixedAsset.objects.create(
        acc_bot_uid=db.AccBot.objects.get(uid=request.POST['fixed_asset_account_uid']),
        asset_no=request.POST['asset_no'],
        asset_name=request.POST['asset_name'],
        get_date=request.POST['get_date'].replace('-',''),
        acquisition_cost=total,
        carrying_value=current_value,
        amortization_way=request.POST['amortization_way'],
        amortization_term_in_month=terms,
        passed_months=passed_month,
        amortization_cost_per_month=monthly_cost,
        amortizated_total_cost=total-current_value,
        sales_income=0,
        is_using=using,
        note=request.POST['note'],
    )

    return redirect('/magi/sdss/fixed_asset/addform/?message=' + message)


def getCostPerMonth(total, terms):
    if int(terms) == 0:
        return 0
    return int(total) / int(terms)


def getPassedMonths(dt_date):
    #ymd = str_date.split('/')
    #if len(ymd) < 3: return 0
    #past_date = datetime(int(ymd[0]),int(ymd[1]),int(ymd[2]))
    if isinstance(dt_date, datetime):
        now_date=u.get_nowdt()
        passed_years=now_date.year-dt_date.year
        passed_months=now_date.month-dt_date.month
        return passed_years*12+passed_months
    else:
        return 0


def getCurrentValue(str_total_cost, str_monthly_cost, str_get_date):

    total_cost = int(str_total_cost)
    monthly_cost = int(str_monthly_cost)
    get_date=getDateTimeFromStr(str_get_date)
    if get_date is not None:
        return total_cost - monthly_cost * getPassedMonths(get_date)
    else:
        return 0


def getDateTimeFromStr(str_date):
    # TODO スクリーニング不十分
    if str_date is not None:
        ymd = str_date.split('-')
        if len(ymd) < 3:
            return None
        return datetime(int(ymd[0]),int(ymd[1]),int(ymd[2]))
    else:
        return None


def getFixedAssetList():
    qs_fixed_asset_list = db.FixedAsset.objects.all()
    d=[]
    for entry in qs_fixed_asset_list:
        d.append({
            'uid': entry.uid,
            'acc_name': entry.acc_bot_uid.name,
            'asset_no': entry.asset_no,
            'asset_name': entry.asset_name,
            'get_date': entry.get_date,
            'acquisition_cost': entry.acquisition_cost,
            'carrying_value': entry.carrying_value,
            'amortization_way': entry.amortization_way,
            'amortization_term_in_month': entry.amortization_term_in_month,
            'passed_months': entry.passed_months,
            'amortization_cost_per_month': entry.amortization_cost_per_month,
            'amortizated_total_cost': entry.amortizated_total_cost,
            'sales_income': entry.sales_income,
            'is_using': entry.is_using,
            'note': entry.note,
        })
    return d


def regularly_view(request):
    # output account_name(str), payment_day(str), amount(int), note(str)
    # toriaezu view only...
    context = {
        'today_date': u.get_nowdt().strftime('%Y-%m-%d'),
        'title_jp': '毎月の支払項目とその金額の一覧です。',
        'regularly_payment_list': getRegularlyPaymentList(),
    }
    return render(request, 'regularly_payment.html', context)


def acc_list(request):
    context = {
        'title_jp': '登録済み勘定科目の一覧です。',
        'acc_list': getAccViewList(),
    }
    return render(request, 'acc_list.html', context)


def getAccViewList():
    qs_top = db.AccTop.objects.all()
    t = {}

    for top_query in qs_top:
        m = {}
        qs_mid = db.AccMid.objects.filter(acc_top_uid = top_query.uid)
        for mid_query in qs_mid:
            b = []
            qs_bot = db.AccBot.objects.filter(acc_mid_uid = mid_query.uid)
            for bot_entry in qs_bot:
                b.append({
                    'uid': bot_entry.uid,
                    'name': bot_entry.name,
                    'note': bot_entry.note,
                })
            m[mid_query.name] = b
        t[top_query.name] = m
    return t


def getRegularlyPaymentList():
    qs = db.RegularlyPayment.objects.all()
    d = []
    for entry in qs:
        d.append({
            'id': entry.uid,
            'is_regist_automaticaly': entry.is_regist_automaticaly,
            'acc_name': entry.acc_bot_uid.name,
            'acc_name_from': entry.acc_bot_uid_from.name,
            'amount_per_month': entry.amount_per_month,
            'payment_day': entry.payment_day,
            'note': entry.note,
        })
    return d


def regularly_add(request):
    context = {
        'title_jp': '毎月の支払項目を登録します。',
        'acc_list': getAccountListByGroup(),
    }
    return render(request, 'regularly_payment_add.html', context)


def regularly_regist(request):

    ret_value = ''
    message=''
    db.RegularlyPayment.objects.create(
        is_regist_automaticaly='is_auto' in request.POST,
        acc_bot_uid=db.AccBot.objects.get(uid=request.POST['acc_bot_uid']),
        acc_bot_uid_from=db.AccBot.objects.get(uid=request.POST['acc_bot_uid_from']),
        amount_per_month=request.POST['amount_per_month'],
        payment_day=request.POST['payment_day'],
        note=request.POST['note'],
    )

    return redirect('/magi/sdss/regularly-payment/add/?message=' + message)

def suii(request, year=0, month=0,
         accID1=GENKIN_ACC_BOT_UID,
         accID2=0,
         accID3=0,
 ):

    message = ''
    # ダミーが来たらとりあえず現金を表示しとく
    if int(accID1) == DUMMY_ACC_BOT_UID:
        accID1 = GENKIN_ACC_BOT_UID
        message += 'accID1にダミーIDが来たため現金を表示します。\r\n'

    year = u.cleanYear(year)
    month = u.cleanMonth(month)
    suiilist1 = getAccountSuii(year, month, accID1)
    suiilist2 = getAccountSuii(year, month, accID2)
    suiilist3 = getAccountSuii(year, month, accID3)
    suiilist = {db.AccBot.objects.get(uid=accID1).name: suiilist1,}
    if int(accID2) != 0 and int(accID2) != DUMMY_ACC_BOT_UID:
        suiilist[db.AccBot.objects.get(uid=accID2).name] = getAccountSuii(year, month, accID2)
    if int(accID3) != 0 and int(accID3) != DUMMY_ACC_BOT_UID:
        suiilist[db.AccBot.objects.get(uid=accID3).name] = getAccountSuii(year, month, accID3)
    labels = list(range(1,32))
    accbot_listgroup = getAccountListByGroup()
    one_two_three = [1,2,3]
    print('are---')
    context = {
        'title_jp':     str(year) + '年' + str(month) + '月の各勘定合計金額変動/累積状況',
        'o_t_t':        one_two_three,
        'listgroup':    accbot_listgroup,
        'ym_list':      getGatherYearMonth(),
        'suii_list':    suiilist,
        'label':        labels,
        'view_name': 'sdss 2.0 Account Transition view',
        'target_year_month': str(year) + '/' + str(month),
        'message': message,
    }
    return render(request, 'suii.html', context)


def getGatherYearMonth():
    qs = db.Journal.objects.annotate(ll=Left('date',6)).values('ll').annotate(Count('ll')).order_by('-ll')
    ym_list = []
    for q in qs:
        if len(q['ll']) >= 6:
            ym_list.append(q['ll'][0:4] + '/' + q['ll'][4:6])
    return ym_list


# TODO dicじゃなくてlistでは・・・・。
def getAccountSuii(year, month, accID):

    if accID == 0:
        return []

    strDate = u.createCurrentYearMonthString(year, month)
    br_direction = db.AccBot.objects.get(uid=accID).acc_mid_uid.acc_top_uid.is_br

    # BSのみ前月末残高を計算
    if db.AccBot.objects.get(uid=accID).acc_mid_uid.acc_top_uid.uid >= 4:
        total = 0
    else:
        init_qs = db.Journal.objects.filter(date__lt=strDate)
        init_br = u.getEmptyOrValueInt(init_qs.filter(br_acc_bot_uid=accID).aggregate(Sum('br_amount'))['br_amount__sum'])
        init_cr = u.getEmptyOrValueInt(init_qs.filter(cr_acc_bot_uid=accID).aggregate(Sum('cr_amount'))['cr_amount__sum'])
        total = u.diffBrCr(init_br, init_cr, br_direction)
    print('total=' + str(total))
    # summing day by day ...
    dic = []
    nextMonth = u.createNextYearMonthString(year, month)
    qs = db.Journal.objects.filter(date__range=(strDate, nextMonth))
    curr_br_qs = qs.filter(br_acc_bot_uid=accID)
    curr_cr_qs = qs.filter(cr_acc_bot_uid=accID)

    # TODO 2019年現在にyearが２０２０だった場合の考慮はしてないが不要？
    isSyoribiYM = u.get_nowdt().year == year and u.get_nowdt().month == month

    # 1日 ~ 31日まで借方と貸方の差分を取得する。2月31日とかも処理するけど文字列型なので問題なし
    for i in range(1,32):
        # 処理当日以降のは登録しない
        if isSyoribiYM and i > u.get_nowdt().day:
            dic.append(0)
        else:
            today_br_sum = u.getEmptyOrValueInt(curr_br_qs.filter(date=(strDate+'{:02}'.format(i))).aggregate(Sum('br_amount'))['br_amount__sum'])
            today_cr_sum = u.getEmptyOrValueInt(curr_cr_qs.filter(date=(strDate+'{:02}'.format(i))).aggregate(Sum('cr_amount'))['cr_amount__sum'])
            print(str(i) + ': add to ' + str(u.diffBrCr(today_br_sum, today_cr_sum, br_direction)))
            total += u.diffBrCr(today_br_sum, today_cr_sum, br_direction)
            dic.append(total)
        #else:
        #    dic.append(0)
    #print(str(dic))
    return dic


def test(request):

    list = getAccountMidAggregateInYear(u.ReportType['PL'])

    context = {
        'pl_list': list,
        'month_list': getMonthList(),
        'view_name': 'sdss 2.0 PL view',
        'target_year_month': '2018/10',
        'message': '',
    }
    return render(request, 'pl.html', context)

def getMonthList():
    month_list = []
    for i in range(12):
        month_list.append('{:02}'.format(i+1))
    return month_list

def getMonthListAndTotal():
    month_total_list = []
    for i in range(12):
        month_total_list.append('{:02}'.format(i+1))
    month_total_list.append('合計')
    return month_total_list



def regist(request):
    #入力チェックと登録用データ作成
    #TODO 入力チェックとUI側の入力チェックもこっちへ移す
    groupid = u.get_nowdt().strftime('%Y%m%d%H%M%S%f')
    strdate = request.POST['journal_date'].replace('-','')
    log.info('regist date as: ' + strdate)

    register_corresp = []
    register_corresp.append({'br_a': 'br_1_a', 'cr_a': 'cr_1_a', 'br_c': 'br_1_c', 'cr_c': 'cr_1_c'})
    register_corresp.append({'br_a': 'br_2_a', 'cr_a': 'cr_2_a', 'br_c': 'br_2_c', 'cr_c': 'cr_2_c'})
    register_corresp.append({'br_a': 'br_3_a', 'cr_a': 'cr_3_a', 'br_c': 'br_3_c', 'cr_c': 'cr_3_c'})
    register_corresp.append({'br_a': 'br_4_a', 'cr_a': 'cr_4_a', 'br_c': 'br_4_c', 'cr_c': 'cr_4_c'})
    register_corresp.append({'br_a': 'br_5_a', 'cr_a': 'cr_5_a', 'br_c': 'br_5_c', 'cr_c': 'cr_5_c'})
    register_corresp.append({'br_a': 'br_6_a', 'cr_a': 'cr_6_a', 'br_c': 'br_6_c', 'cr_c': 'cr_6_c'})
    register_corresp.append({'br_a': 'br_7_a', 'cr_a': 'cr_7_a', 'br_c': 'br_7_c', 'cr_c': 'cr_7_c'})
    register_corresp.append({'br_a': 'br_8_a', 'cr_a': 'cr_8_a', 'br_c': 'br_8_c', 'cr_c': 'cr_8_c'})
    registerd = False
    note_temp = ''
    for cor in register_corresp:
        if u.isIntAndNotZero(request.POST[cor['br_a']]) or u.isIntAndNotZero(request.POST[cor['cr_a']]):
            if registerd == True:
                note_temp = ''
            else:
                note_temp = request.POST['note']
            db.Journal.objects.create(
                date = strdate,
                group_id = groupid,
                br_acc_bot_uid = db.AccBot.objects.get(uid=request.POST[cor['br_c']]),
                br_amount = u.getEmptyOrValueInt(request.POST[cor['br_a']]),
                cr_acc_bot_uid = db.AccBot.objects.get(uid=request.POST[cor['cr_c']]),
                cr_amount = u.getEmptyOrValueInt(request.POST[cor['cr_a']]),
                note = note_temp,
            )
            registerd = True
            log.info('register journal object')
    # TODO ハードコーディングなのでショートカットにするとか200
    href = "{% url 'journal' %}"
    #return redirect('/magi/sdss')
    return redirect(reverse('sdss_index'))


def regist_regularly_payment(request):
    '''指定された年月日に登録済みの定期支払項目を一括で自動登録する'''
    today_str = u.createCurrentDateString()
    strdate = request.POST['reg_regist_date'].replace('-', '')
    note = u.get_nowdt().strftime('%Y/%m/%d %H:%M') + ' 自動登録です'
    qs_reg_list = db.RegularlyPayment.objects.all()
    groupid = u.get_nowdt().strftime('%Y%m%d%H%M%S%f')
    for qs in qs_reg_list:
        if qs.is_regist_automaticaly == 1:
            db.Journal.objects.create(
                date = strdate,
                group_id = groupid,
                br_acc_bot_uid = db.AccBot.objects.get(uid=qs.acc_bot_uid.uid),
                br_amount = u.getEmptyOrValueInt(qs.amount_per_month),
                cr_acc_bot_uid = db.AccBot.objects.get(uid=qs.acc_bot_uid_from.uid),
                cr_amount = u.getEmptyOrValueInt(qs.amount_per_month),
                note = note,
            )
    #return redirect('/magi/sdss/regularly-payment/')
    return redirect(reverse('regularly_view'))


import openpyxl


def journal_export(request, year=0):
    '''指定された年の1〜12月の仕訳を記載したエクセルファイルをダウンロードする'''
    yaer = u.cleanYear(year)
    filename = 'fs_'
    filename += year
    filename += '_'
    filename += u.getStrTimeStamp()
    filename += '.xlsx'
    wb = openpyxl.load_workbook('/home/share/template_f_s.xlsx')
    y_startDate = year + '0101'
    y_endDate = year + '1231'
    journal_qs = db.Journal.objects.filter(date__range=(y_startDate,y_endDate)).order_by('date')
    journal_list = getExportJournalList(journal_qs)
    currentSheetName = '01'
    sheet = wb[currentSheetName]
    current_col=1
    current_row=5
    for journal in journal_list:
        if currentSheetName != journal['month']:
            currentSheetName = journal['month']
            sheet = wb[currentSheetName]
            current_row=5
        sheet.cell(row=current_row, column=1, value=journal['month'])
        sheet.cell(row=current_row, column=2, value=journal['day'])
        sheet.cell(row=current_row, column=3, value=journal['br_name'])
        sheet.cell(row=current_row, column=4, value=journal['br_amount'])
        sheet.cell(row=current_row, column=5, value=journal['cr_name'])
        sheet.cell(row=current_row, column=6, value=journal['cr_amount'])
        sheet.cell(row=current_row, column=7, value=journal['note'])
        current_row += 1

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    wb.save(response)
    return response


def journal(request):
    # | uid | date | group_id | br_acc_bot_name | br_amount | ... |
    # TODO extract year, month ...etc
    journal_qs = db.Journal.objects.order_by('-date')
    journal_list = getJournalList(journal_qs)
    # yearList = ['2019', '2018',]
    yearList = u.getSelectableYearList()
    context = {
        'journal_list': journal_list,
        'year_list':    yearList,
        'view_name':    'sdss 2.0 journal view',''
        'message': '',
    }
    return render(request, 'journal.html', context)


def asset_suii(request, year=0):
    year = u.cleanYear(year)
    bs = b()
    bslist = bs.getMiddleYearStatement(year, 0)
    bslist.pop('自己資本', 'no match')
    bslist.pop('他人資本', 'no match')
    bslist.pop('評価換算差額等', 'no match')
    bslist.pop('固定資産', 'no match')
    # yearList = ['2019', '2018',]
    yearList = u.getSelectableYearList()
    monthList = getMonthList()
    colors = (
        (0,153,255,10),
        (0,204,204,10),
        (0.204,153,10),
        (0,0,102,10),
        (0,51,51,10),
        (51,0,51,10),
        (102,0,0,10),
        (102,102,0,10),
        (255,102,0,10),
    )
    context = {
        'label': monthList,
        'bs_list': bslist,
        'year_list': yearList,
        'colors': colors,
        'view_name': 'sdss 2.0 Asset Summary Suii',
        'target_year': str(year),
        'message': '',
    }
    return render(request, 'asset_suii.html', context)


def summary(request, year=0):
    year = u.cleanYear(year)
    bs = b()
    pl = p()
    bslist = bs.getMiddleYearStatement(year)
    pllist = pl.getMiddleYearStatement(year)
    # yearList = ['2019', '2018',]
    yearList = u.getSelectableYearList()
    context = {
        'bs_list': bslist,
        'pl_list': pllist,
        'month_list': getMonthList(),
        'month_list_total': getMonthListAndTotal(),
        'year_list': yearList,
        'view_name': 'sdss 2.0 PL BS summary view',
        'target_year': str(year),
        'message': '',
    }
    return render(request, 'summary.html', context)


def bs(request, year=0):
    # |       |   10    |    11   |    12   |
    # | ICOCA |   2,100 |   5,900 |   9,200 |
    # |  現金  |  13,110 |  24,670 |     ... |
    year = u.cleanYear(year)
    bs_class = b()
    bslist = bs_class.getBottomYearStatement(year)
    # yearList = ['2019', '2018',]
    yearList = u.getSelectableYearList()
    context = {
        'bs_list': list,
        'year_list': yearList,
        'bs_mid_list': bslist,
        'month_list': getMonthList(),
        'view_name': 'sdss 2.0 BS view',
        'target_year': str(year),
        'message': '',
    }
    return render(request, 'bs.html', context)


def pl(request, year=0):
    year = u.cleanYear(year)
    pl = p()
    pllist = pl.getBottomYearStatement(year)
    # yearList = ['2019', '2018',]
    yearList = u.getSelectableYearList()
    context = {
        'year_list': yearList,
        'pl_list': pllist,
        #'month_list': getMonthList(),
        'month_list': getMonthListAndTotal(),
        'view_name': 'sdss 2.0 PL view',
        'target_year': str(year),
        'message': '',
    }
    return render(request, 'pl.html', context)


def rigid_pl(request, year=0):
    year = u.cleanYear(year)
    r_pl = r_p()

    #売上総利益
    netRevenue = r_pl.getNetRevenue(year)
    costOfSales = r_pl.getCostOfSales(year)
    grossMarginList = []
    for i in range(12):
        grossMarginList.append(netRevenue['total'][i] - costOfSales['total'][i])
    grossMargin = {
        '収益': netRevenue,
        '売上原価': costOfSales,
        '売上総利益': grossMarginList,
    }
    #
    # #限界利益
    # taxAndFocedCost = r_pl.getTaxAndForcedCost(year)
    # lifeCost = r_pl.getLifeCost(year)
    # limitIncomeList = []
    # for i in range(12):
    #     limitIncomeList.append(grossMarginList[i] - taxAndFocedCost[i] - lifeCost[i])
    # limitIncome = {
    #     '一般費':  taxAndFocedCost,
    #     '生活費':  lifeCost,
    #     '限界利益':limitIncomeList,
    # }
    #
    # #その他の収益・費用等
    # nonOperatingCost = r_pl.getNonOperatingCost(year)
    # nonOperatingIncome = r_pl.getNonOperatingIncome(year)
    # entertainmentCost = r_pl.getEntertainmentCost(year)
    #
    # netIncomeList = []
    # for i in range(12):
    #     netIncomeList.append(
    #             limitIncomeList[i]
    #          -  nonOperatingCost[i]
    #          +  nonOperatingIncome[i]
    #          -  entertainmentCost[i]
    #     )
    # netIncome = {
    #     '営業外収益': nonOperatingIncome,
    #     '営業外費用': nonOperatingCost,
    #     '娯楽費': entertainmentCost,
    #     '当月純利益': netIncomeList,
    # }

    yearList = u.getSelectableYearList()
    context = {
        'year_list': yearList,
        '売上総利益': grossMargin,
        # '限界利益': limitIncome,
        # '当月純利益': netIncome,
        # 'month_list': getMonthList(),
        # 'view_name': 'sdss 2.0 Rigid PL view',
        # 'target_year': str(year),
        # 'message': '',
    }
    return render(request, 'rigid_pl.html', context)


def detail(request, mid_class_uid=1, year=0):
    year = u.cleanYear(year)
    d_class = d()
    # selectable_year_list = ['2019', '2018',]
    selectable_year_list = u.getSelectableYearList()
    month_list = range(1,13) # =['1','2', .. '12']
    mid_class_list = d_class.getTopMidAccList()
    acc_amount_list = d_class.getClassificationDetail(mid_class_uid, year)
    mid_class_name = d_class.getMidClassName(mid_class_uid)
    context = {
        'year': year,
        'month_list': month_list,
        'selectable_year_list': selectable_year_list,
        'mid_class_name': mid_class_name,
        'mid_class_list': mid_class_list,
        'acc_amount_list': acc_amount_list,
    }
    return render(request, 'detail.html', context)


def cs(request):
    raise NotImplementedError()


def getUidAndName(qs_accbot):
    d = {}
    for entry in qs_accbot:
        if entry.uid in d:
            raise Error('duplicate uid=' + entry.uid + ' in AccBot ')
        else:
            d[entry.uid] = entry.name
    return d


def getAccountList(qs_accbot):
    d = []
    for entry in qs_accbot:
        d.append({
            'uid':    entry.uid,
            'label':  entry.name_ac,
            'value':  entry.name,
        })
    return d


def getAccountListByGroup():
    qs_bot = db.AccBot.objects.order_by('sort_order').all()
    d = {}
    for entry in qs_bot:
        if not entry.acc_mid_uid.name in d:
            d[entry.acc_mid_uid.name] = []

        d[entry.acc_mid_uid.name].append({
            'uid':    entry.uid,
            'label':  entry.name_ac,
            'value':  entry.name,
        })
    return d


def getJournalList(qs_journal):
    d = []
    for entry in qs_journal:
        d.append({
            'id':        entry.uid,
            'date':      entry.date[:4] + '/' + entry.date[4:6] + '/'+ entry.date[6:8],
            'group_id':  entry.group_id,
            'br_name':   entry.br_acc_bot_uid.name if entry.br_acc_bot_uid.name != '（選択してください）' else '',
            'br_amount': entry.br_amount if entry.br_amount != 0 else '',
            'cr_name':   entry.cr_acc_bot_uid.name if entry.cr_acc_bot_uid.name != '（選択してください）' else '',
            'cr_amount': entry.cr_amount if entry.cr_amount != 0 else '',
            'note':      entry.note,
        })
    return d

def getExportJournalList(qs_journal):
    d = []
    for entry in qs_journal:
        br_name = ''
        cr_name =''
        if entry.br_acc_bot_uid.name == '（選択してください）':
            br_name = ''
        elif entry.br_acc_bot_uid.export_name != 'null':
            br_name = entry.br_acc_bot_uid.export_name
        else:
            br_name = entry.br_acc_bot_uid.name

        if entry.cr_acc_bot_uid.name == '（選択してください）':
            cr_name = ''
        elif entry.cr_acc_bot_uid.export_name  != 'null':
            cr_name = entry.cr_acc_bot_uid.export_name
        else:
            cr_name = entry.cr_acc_bot_uid.name

        d.append({
            'id':        entry.uid,
            'year':      entry.date[:4],
            'month':     entry.date[4:6],
            'day':       entry.date[6:8],
            'group_id':  entry.group_id,
            'br_name':   br_name,
            'br_amount': entry.br_amount if entry.br_amount != 0 else '',
            'cr_name':   cr_name,
            'cr_amount': entry.cr_amount if entry.cr_amount != 0 else '',
            'note':      entry.note,
        })
    return d


def wants_list(request):
    qs_bot = db.WantsList.objects.all()
    d = []
    for entry in qs_bot:
        d.append({
            'id': entry.uid,
            'name': entry.name,
            'date': entry.date,
            'amount': entry.amount,
            'note': entry.note,
        })
    context = {
        'view_name': 'ほしいものリスト',
        # 'year': year,
        # 'selectable_year_list': selectable_year_list,
        'wants_list': d,
    }
    return render(request, 'wants_list.html', context)


#TODO すでに処理が重そうな気がするので今後パフォーマンスに関する何らかの処置が必要かも
#TODO 貸借差額=当期損益の計算および表示
